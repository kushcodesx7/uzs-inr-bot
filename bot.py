import html
import json
import os
import statistics
import sys
import traceback
from datetime import datetime, timedelta, timezone
from pathlib import Path
from zoneinfo import ZoneInfo

import requests
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

STATE_FILE = Path("last_rate.json")
HISTORY_FILE = Path("history.xlsx")
DASHBOARD_DATA = Path("docs/data.json")
TZ = ZoneInfo("Asia/Tashkent")
ALERT_THRESHOLD_INR = 200.0
RATE_SOURCES = [
    # Wise live rate — near-real-time (few-minute lag), same feed powering
    # Wise's own UZS→INR widget, which is what Google's converter mirrors.
    {
        "name": "wise",
        "url": "https://wise.com/rates/live?source=UZS&target=INR",
        "parser": lambda d: d["value"],
        "headers": {"User-Agent": "Mozilla/5.0 (compatible; UZS-INR-Tracker)"},
    },
    # open.er-api.com — hourly update, free, no key.
    {
        "name": "er-api",
        "url": "https://open.er-api.com/v6/latest/UZS",
        "parser": lambda d: d["rates"]["INR"],
        "headers": None,
    },
    # fawazahmed0/currency-api via jsDelivr — daily aggregate, reliable backup.
    {
        "name": "fawazahmed0-jsdelivr",
        "url": "https://cdn.jsdelivr.net/npm/@fawazahmed0/currency-api@latest/v1/currencies/uzs.json",
        "parser": lambda d: d["uzs"]["inr"],
        "headers": None,
    },
    # Same data via pages.dev CDN — last-resort fallback.
    {
        "name": "fawazahmed0-pages",
        "url": "https://latest.currency-api.pages.dev/v1/currencies/uzs.json",
        "parser": lambda d: d["uzs"]["inr"],
        "headers": None,
    },
]
TELEGRAM_API = "https://api.telegram.org/bot{token}/sendMessage"
CRON_INTERVAL_MIN = 30
DIVIDER = "━━━━━━━━━━━━━━━━"
CHECKS_PER_DAY = (24 * 60) // CRON_INTERVAL_MIN

HEADERS = [
    "Date",
    "Time",
    "Rate (1 UZS = INR)",
    "INR amount",
    "Change (INR)",
    "% Change",
    "Direction",
]

GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")


def format_indian(n, decimals=2):
    sign = "-" if n < 0 else ""
    n = abs(float(n))
    if decimals == 0:
        int_part = f"{round(n):d}"
        dec_part = ""
    else:
        int_part, dec_part = f"{n:.{decimals}f}".split(".")
    if len(int_part) <= 3:
        out = f"{sign}{int_part}"
    else:
        last3 = int_part[-3:]
        rest = int_part[:-3]
        groups = []
        while len(rest) > 2:
            groups.insert(0, rest[-2:])
            rest = rest[:-2]
        if rest:
            groups.insert(0, rest)
        out = f"{sign}{','.join(groups)},{last3}"
    return f"{out}.{dec_part}" if dec_part else out


def build_history_table(records, limit=12):
    """Render the last N checks as a date-grouped diary <pre> block with
    🟢/🔴/⚪ markers. Most recent date first; within a day, newest check on top.
    The very latest row is tagged "← now"."""
    if not records:
        return ""
    recent = records[-limit:]
    latest = recent[-1]

    groups = {}
    order = []
    for r in recent:
        d = r.get("date", "")
        if d not in groups:
            groups[d] = []
            order.append(d)
        groups[d].append(r)

    lines = []
    for i, date_str in enumerate(reversed(order)):
        try:
            dt = datetime.strptime(date_str, "%Y-%m-%d")
            header = dt.strftime("%d %b").lstrip("0")
        except ValueError:
            header = date_str
        if i > 0:
            lines.append("")
        lines.append(f"━ {header} ━")
        for r in reversed(groups[date_str]):
            t = (r.get("time") or "")[:5]
            inr_str = f"₹{format_indian(r.get('inr', 0), 0)}"
            change = r.get("change") or 0
            direction = r.get("direction", "")
            if direction == "START":
                marker, delta_str = "⚪", "start"
            elif direction == "UP":
                marker, delta_str = "🟢", f"+{format_indian(abs(change), 0)}"
            elif direction == "DOWN":
                marker, delta_str = "🔴", f"−{format_indian(abs(change), 0)}"
            else:
                marker, delta_str = "⚪", "flat"
            is_now = (r is latest)
            suffix = "   ← now" if is_now else ""
            lines.append(f"  {t}   {inr_str:<11}  {marker} {delta_str}{suffix}")
    return "<pre>" + "\n".join(lines) + "</pre>"


def fetch_rate():
    last_err = None
    for src in RATE_SOURCES:
        try:
            resp = requests.get(src["url"], timeout=30, headers=src.get("headers") or {})
            resp.raise_for_status()
            rate = float(src["parser"](resp.json()))
            if rate <= 0:
                raise RuntimeError(f"non-positive rate {rate}")
            print(f"[rate] source={src['name']} value={rate}", file=sys.stderr)
            return rate, src["name"]
        except Exception as e:
            last_err = e
            print(f"[rate] {src['name']} failed: {e}", file=sys.stderr)
    raise RuntimeError(f"All rate sources failed: {last_err}")


def load_state():
    if not STATE_FILE.exists():
        return None
    return json.loads(STATE_FILE.read_text())


def save_state(state):
    STATE_FILE.write_text(json.dumps(state, indent=2))


def open_workbook():
    if HISTORY_FILE.exists():
        return load_workbook(HISTORY_FILE)
    wb = Workbook()
    ws = wb.active
    ws.title = "History"
    ws.append(HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    widths = [12, 10, 20, 16, 14, 10, 12]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    return wb


def log_row(wb, date_str, time_str, rate, inr_amount, change, pct_change, direction):
    ws = wb.active
    ws.append([date_str, time_str, rate, inr_amount, change, pct_change, direction])
    row = ws.max_row
    ws.cell(row=row, column=3).number_format = "0.00000000"
    ws.cell(row=row, column=4).number_format = "#,##0.00"
    ws.cell(row=row, column=5).number_format = "#,##0.00"
    ws.cell(row=row, column=6).number_format = "0.00"
    direction_cell = ws.cell(row=row, column=7)
    if direction == "UP":
        direction_cell.fill = GREEN_FILL
    elif direction == "DOWN":
        direction_cell.fill = RED_FILL
    wb.save(HISTORY_FILE)


def todays_range(current_inr, today_str):
    high = low = current_inr
    if not HISTORY_FILE.exists():
        return high, low
    wb = load_workbook(HISTORY_FILE, read_only=True)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 4:
            continue
        date_val, inr_val = row[0], row[3]
        if date_val == today_str and isinstance(inr_val, (int, float)):
            high = max(high, inr_val)
            low = min(low, inr_val)
    wb.close()
    return high, low


def read_history():
    if not HISTORY_FILE.exists():
        return []
    wb = load_workbook(HISTORY_FILE, read_only=True)
    ws = wb.active
    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        date_v, time_v, rate_v, inr_v, change_v, pct_v, direction_v = (row + (None,) * 7)[:7]
        if not isinstance(inr_v, (int, float)):
            continue
        records.append({
            "date": str(date_v),
            "time": str(time_v),
            "rate": float(rate_v) if isinstance(rate_v, (int, float)) else None,
            "inr": float(inr_v),
            "change": float(change_v) if isinstance(change_v, (int, float)) else 0.0,
            "pct_change": float(pct_v) if isinstance(pct_v, (int, float)) else 0.0,
            "direction": str(direction_v) if direction_v else "",
        })
    wb.close()
    return records


def _linreg_slope(ys):
    n = len(ys)
    if n < 3:
        return None
    xs = list(range(n))
    mx = sum(xs) / n
    my = sum(ys) / n
    num = sum((x - mx) * (y - my) for x, y in zip(xs, ys))
    den = sum((x - mx) ** 2 for x in xs)
    return num / den if den else 0.0


def compute_analytics(records, current_inr):
    inrs = [r["inr"] for r in records if r["inr"] is not None]
    if not inrs:
        return None

    w24 = inrs[-CHECKS_PER_DAY:] if len(inrs) >= 2 else []
    w7d = inrs[-CHECKS_PER_DAY * 7:] if len(inrs) >= 3 else []
    w30d = inrs[-CHECKS_PER_DAY * 30:] if len(inrs) >= 5 else []

    ma_24h = statistics.mean(w24) if len(w24) >= 2 else None
    ma_7d = statistics.mean(w7d) if len(w7d) >= 3 else None
    ma_30d = statistics.mean(w30d) if len(w30d) >= 5 else None
    stddev_7d = statistics.pstdev(w7d) if len(w7d) >= 3 else None
    slope_7d = _linreg_slope(w7d) if len(w7d) >= 3 else None

    zscore = None
    if ma_7d is not None and stddev_7d and stddev_7d > 0:
        zscore = (current_inr - ma_7d) / stddev_7d

    high_30d = max(w30d) if w30d else None
    low_30d = min(w30d) if w30d else None
    pct_from_high = ((current_inr - high_30d) / high_30d * 100) if high_30d else None
    pct_from_low = ((current_inr - low_30d) / low_30d * 100) if low_30d else None

    advisory = "Not enough history yet — tracking will improve signals over the next few days."
    level = "neutral"
    if zscore is not None and slope_7d is not None:
        trending_up = slope_7d > 0
        if zscore >= 1 and trending_up:
            advisory = "ABOVE avg & rising — relatively favorable to convert UZS→INR; watch for reversal."
            level = "convert"
        elif zscore >= 1 and not trending_up:
            advisory = "ABOVE avg but weakening — consider converting before it mean-reverts."
            level = "convert-soft"
        elif zscore <= -1 and trending_up:
            advisory = "BELOW avg but turning up — patience may pay off, trend is improving."
            level = "hold"
        elif zscore <= -1 and not trending_up:
            advisory = "BELOW avg & still falling — holding is risky; downtrend not exhausted."
            level = "hold-risky"
        else:
            advisory = "Near recent average — no strong signal either way."
            level = "neutral"

    return {
        "current_inr": current_inr,
        "ma_24h": ma_24h,
        "ma_7d": ma_7d,
        "ma_30d": ma_30d,
        "stddev_7d": stddev_7d,
        "slope_7d": slope_7d,
        "zscore": zscore,
        "high_30d": high_30d,
        "low_30d": low_30d,
        "pct_from_high_30d": pct_from_high,
        "pct_from_low_30d": pct_from_low,
        "advisory": advisory,
        "level": level,
        "data_points": len(inrs),
    }


def write_dashboard_data(records, analytics, current_rate, now):
    DASHBOARD_DATA.parent.mkdir(parents=True, exist_ok=True)
    payload = {
        "updated_iso": now.isoformat(),
        "updated_display": now.strftime("%d %b %Y, %-I:%M %p %Z"),
        "alert_threshold_inr": ALERT_THRESHOLD_INR,
        "current_rate": current_rate,
        "analytics": analytics,
        "history": records,
    }
    DASHBOARD_DATA.write_text(json.dumps(payload, indent=2, default=str))


def next_check_display(now, tz):
    now_utc = now.astimezone(timezone.utc).replace(second=0, microsecond=0)
    minutes_to_add = CRON_INTERVAL_MIN - (now_utc.minute % CRON_INTERVAL_MIN)
    next_run = now_utc + timedelta(minutes=minutes_to_add)
    return next_run.astimezone(tz).strftime("%-I:%M %p")


def send_telegram(token, chat_id, text):
    url = TELEGRAM_API.format(token=token)
    resp = requests.post(
        url,
        json={
            "chat_id": chat_id,
            "text": text,
            "parse_mode": "HTML",
            "disable_web_page_preview": True,
        },
        timeout=30,
    )
    resp.raise_for_status()
    return resp.json()


def build_message(inr_amount, change, direction, records):
    """Minimal alert: one-line headline + date-grouped diary of recent checks."""
    amount_str = format_indian(inr_amount, 0)
    if direction == "START":
        headline = f"💱 <b>₹{amount_str}</b>  ·  baseline set"
    elif direction == "UP":
        headline = f"💱 <b>₹{amount_str}</b>   🟢 +₹{format_indian(abs(change), 0)}"
    elif direction == "DOWN":
        headline = f"💱 <b>₹{amount_str}</b>   🔴 −₹{format_indian(abs(change), 0)}"
    else:
        headline = f"💱 <b>₹{amount_str}</b>"

    table = build_history_table(records, limit=12)
    return headline + ("\n\n" + table if table else "")


def _run_tracker():
    try:
        token = os.environ["BOT_TOKEN"]
        chat_id = os.environ["CHAT_ID"]
        amount_uzs = float(os.environ["AMOUNT_UZS"])
    except KeyError as e:
        print(f"Missing required env var: {e}", file=sys.stderr)
        sys.exit(1)

    rate, source = fetch_rate()
    inr_amount = amount_uzs * rate

    now = datetime.now(TZ)
    date_str = now.strftime("%Y-%m-%d")
    time_str = now.strftime("%H:%M:%S")
    time_display = now.strftime("%-I:%M %p")
    full_display = now.strftime("%Y-%m-%d %H:%M %Z")

    prev = load_state()
    prev_source = (prev or {}).get("source")
    source_switched = bool(prev_source) and prev_source != source

    if prev is None:
        direction = "START"
        change = 0.0
        pct_change = 0.0
    else:
        change = inr_amount - prev["inr_amount"]
        base = prev["inr_amount"] or 1.0
        pct_change = (change / base) * 100
        if change > 0:
            direction = "UP"
        elif change < 0:
            direction = "DOWN"
        else:
            direction = "FLAT"

    wb = open_workbook()
    log_row(wb, date_str, time_str, rate, inr_amount, change, pct_change, direction)

    records = read_history()
    analytics = compute_analytics(records, inr_amount)
    write_dashboard_data(records, analytics, rate, now)

    # Baseline-only alert when the rate source changes, because different
    # sources have slightly different mid-market rates and the numeric
    # delta between them isn't a real market move.
    if source_switched:
        print(f"[{full_display}] Source switched {prev_source} -> {source}; skipping alert to avoid false positive.")
        should_send = False
    else:
        should_send = direction == "START" or abs(change) > ALERT_THRESHOLD_INR

    if should_send:
        msg = build_message(inr_amount, change, direction, records)
        send_telegram(token, chat_id, msg)
        print(f"[{full_display}] Sent ({direction}) via {source}: change={change:+.2f}")
    else:
        print(f"[{full_display}] No alert via {source}: change={change:+.2f} under ₹{ALERT_THRESHOLD_INR:.0f}")

    save_state({
        "rate": rate,
        "inr_amount": inr_amount,
        "timestamp": now.isoformat(),
        "timestamp_display": full_display,
        "timestamp_time": time_display,
        "source": source,
    })


def main():
    try:
        _run_tracker()
    except SystemExit:
        raise
    except Exception:
        # Best-effort: tell the user Telegram-side that the tracker is broken,
        # so silent failures don't go unnoticed. Still exit non-zero afterwards
        # so the workflow surfaces the failure.
        tb = traceback.format_exc()
        print(tb, file=sys.stderr)
        token = os.environ.get("BOT_TOKEN", "")
        chat_id = os.environ.get("CHAT_ID", "")
        if token and chat_id:
            safe_tb = html.escape(tb[-1000:])
            try:
                send_telegram(token, chat_id, f"⚠️ <b>Tracker error</b>\n<pre>{safe_tb}</pre>")
            except Exception as e:
                print(f"Also failed to send error Telegram: {e}", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
